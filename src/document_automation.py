import os
from pypdf import PdfReader
from pypdf.errors import DependencyError, PdfReadError
from openpyxl import Workbook

PDF_FOLDER = r"C:\Users\BrandonEdgerson\OneDrive - CareStar Inc\Carestar\OHCWEAT"
OUTPUT_PATH = r"C:\Users\BrandonEdgerson\OneDrive - CareStar Inc\Carestar\OHCWEAT\Best_Updated_Test_Merge_Sheet.xlsx"

NUM_RECENT_PDFS = 65

MAILMERGE_SHEET = "MailMerge"
MEDS_SHEET = "Medications"

MAIL_HEADERS = [
    "ID", "First Name", "Last Name", "DOB", "Gender", "Phone Number",
    "Physical Address", "City", "State", "Zip", "Medicaid ID", "Language",
    "Cultural Considerations", "CRM Fist Name", "CRM Last name",
    "CRM Phone Number", "CRM Email", "PCP First Name", "PCP Last Name",
    "PCP Phone Number", "PCP Fax Number", "PCP Address", "PCP City",
    "PCP State", "PCP Zip Code", "FFC First Name", "FFC Last Name",
    "FFC Phone Number", "WhoMonitorsMedications", "WhoSetsUpMedication",
    "Goals"
]

MED_HEADERS = [
    "ID", "First Name", "Last Name", "Medication", "Where", "Mail",
    "Dose", "Frequency", "Route", "Packaging", "Assistance",
    "Who Assists", "Prescribed By"
]


def clean(value):
    return "" if value is None else str(value).strip()


def get_field(fields, name):
    field = fields.get(name)
    if not field:
        return ""
    return clean(field.get("/V"))


def get_language(fields):
    languages = [
        "English", "Russian", "Spanish", "Somali", "Nepalese", "Arabic",
        "Serbian", "Laotian", "Cantonese", "Estonian", "Mandarin"
    ]

    selected = []

    for lang in languages:
        field = fields.get(lang)
        if field and field.get("/V"):
            selected.append(lang)

    return ", ".join(selected)


def append_by_headers(ws, data, headers):
    ws.append([data.get(header, "") for header in headers])


def get_next_client_id(ws):
    return f"C{ws.max_row:03d}"


def extract_client(fields, client_id):
    return {
        "ID": client_id,
        "First Name": get_field(fields, "First Name").upper(),
        "Last Name": get_field(fields, "Last Name").upper(),
        "DOB": get_field(fields, "Date of Birth mmddyyyy"),
        "Gender": get_field(fields, "Sex"),
        "Phone Number": get_field(fields, "Telephone primary"),
        "Physical Address": get_field(fields, "Permanent Address"),
        "City": get_field(fields, "City"),
        "State": get_field(fields, "State"),
        "Zip": get_field(fields, "Zip Code"),
        "Medicaid ID": get_field(fields, "Text3"),
        "Language": get_language(fields),
        "Cultural Considerations": get_field(fields, "Note any cultural preferences"),

        "CRM Fist Name": get_field(fields, "Assessor First Name"),
        "CRM Last name": get_field(fields, "Assessor Last Name"),
        "CRM Phone Number": "",
        "CRM Email": "",

        "PCP First Name": get_field(fields, "First Name_4"),
        "PCP Last Name": get_field(fields, "Last Name_4"),
        "PCP Phone Number": get_field(fields, "Telephone primary_2"),
        "PCP Fax Number": get_field(fields, "Fax Number"),
        "PCP Address": get_field(fields, "Address_3"),
        "PCP City": get_field(fields, "City_4"),
        "PCP State": get_field(fields, "State_4"),
        "PCP Zip Code": get_field(fields, "Zip Code_4"),

        "FFC First Name": get_field(fields, "First Name_2"),
        "FFC Last Name": get_field(fields, "Last Name_2"),
        "FFC Phone Number": get_field(fields, "Mobile Telephone"),

        "WhoMonitorsMedications": get_field(
            fields,
            "Who monitors medications specify name and relationship"
        ),
        "WhoSetsUpMedication": get_field(
            fields,
            "Who sets up the medication and how often is this performed"
        ),
        "Goals": get_field(
            fields,
            "Does the individual have a goals Yes No If yes what is the individuals goals"
        ),
    }


def extract_meds(fields, client_id, first_name, last_name):
    meds = []

    for i in range(1, 101):
        medication = get_field(fields, f"MedicationRow{i}")

        if not medication:
            continue

        meds.append({
            "ID": client_id,
            "First Name": first_name,
            "Last Name": last_name,
            "Medication": medication,
            "Where": get_field(fields, f"Where ObtainedRow{i}"),
            "Mail": get_field(fields, f"Mail OrderRow{i}"),
            "Dose": get_field(fields, f"DoseRow{i}"),
            "Frequency": get_field(fields, f"FrequencyRow{i}"),
            "Route": get_field(fields, f"RouteRow{i}"),
            "Packaging": get_field(fields, f"PackagingRow{i}"),
            "Assistance": get_field(fields, f"Assistance NeededRow{i}"),
            "Who Assists": get_field(fields, f"Who AssistsRow{i}"),
            "Prescribed By": get_field(fields, f"Prescribed ByRow{i}"),
        })

    return meds


def get_recent_pdfs(folder, count):
    pdf_paths = [
        os.path.join(folder, file)
        for file in os.listdir(folder)
        if file.lower().endswith(".pdf")
    ]

    pdf_paths.sort(key=os.path.getmtime, reverse=True)

    return pdf_paths[:count]


def main():
    if not os.path.exists(PDF_FOLDER):
        raise FileNotFoundError(f"PDF folder not found: {PDF_FOLDER}")

    wb = Workbook()

    ws_mail = wb.active
    ws_mail.title = MAILMERGE_SHEET
    ws_mail.append(MAIL_HEADERS)

    ws_meds = wb.create_sheet(MEDS_SHEET)
    ws_meds.append(MED_HEADERS)

    pdf_paths = get_recent_pdfs(PDF_FOLDER, NUM_RECENT_PDFS)

    if not pdf_paths:
        print("No PDF files found.")
        return

    print("Processing most recent PDFs:")

    for pdf_path in pdf_paths:
        pdf_file = os.path.basename(pdf_path)
        print(f"Reading: {pdf_file}")

        try:
            reader = PdfReader(pdf_path)
            fields = reader.get_fields()

        except DependencyError:
            print(f"SKIPPED encrypted PDF needing cryptography: {pdf_file}")
            continue

        except PdfReadError:
            print(f"SKIPPED unreadable PDF: {pdf_file}")
            continue

        except Exception as e:
            print(f"SKIPPED error with {pdf_file}: {e}")
            continue

        if not fields:
            print(f"SKIPPED no fillable form fields: {pdf_file}")
            continue

        client_id = get_next_client_id(ws_mail)
        client = extract_client(fields, client_id)

        if not client["First Name"] and not client["Last Name"]:
            print(f"SKIPPED no client name found: {pdf_file}")
            continue

        append_by_headers(ws_mail, client, MAIL_HEADERS)

        meds = extract_meds(
            fields,
            client_id,
            client["First Name"],
            client["Last Name"]
        )

        for med in meds:
            append_by_headers(ws_meds, med, MED_HEADERS)

        print(
            f"Added {client_id}: "
            f"{client['First Name']} {client['Last Name']} | "
            f"Meds: {len(meds)} | "
            f"Goal: {client['Goals']}"
        )

    wb.save(OUTPUT_PATH)

    print("DONE")
    print(f"Saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
